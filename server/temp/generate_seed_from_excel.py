from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
EXCEL = ROOT / 'SL_Keys-AI.xlsx'
OUT = ROOT / 'server' / 'src' / 'database' / 'seeds' / 'seed-current-prices.sql'

wb = load_workbook(EXCEL, data_only=True)
ws = wb[wb.sheetnames[0]]

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
idx = {h: i for i, h in enumerate(headers)}

required = [
    'YEAR',
    'MAKE',
    'MODEL',
    'TYPE_OF_IGNITION',
    'TYPE_OF_KEY',
    'NO_BUTTONS',
    'SL_TRANSMITTER_PRICE ',
    'SL_ADD_KEY_PRICE',
    'SL_PROGRAM_PRICE',
    'DEALER_TRANSMITTER_PRICE',
    'DEALER_EMERGENCY_BLADE_PRICE',
    'DEALER_PROGRAM_PRICE',
    'DEALER_NAME',
    'DEALER_PHONE',
    'DEALER_DATE_CALLED',
    'DEALER_ADDRESS',
    'DEALER_VIN',
    'DEALER_PART_NUMBER',
    'LINK',
    'COMMENTS',
]

for key in required:
    if key not in idx:
        raise RuntimeError(f'Missing expected column: {key}')


def clean_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def collapse_spaces(v):
    if v is None:
        return None
    return re.sub(r'\s+', ' ', v).strip()


def normalize_make(v):
    s = collapse_spaces(v)
    if not s:
        return None

    key = s.lower().replace('-', ' ')
    key = re.sub(r'\s+', ' ', key).strip()

    if key in {'mercedes', 'mercedes benz'}:
        return 'mercedes'
    if key == 'kia':
        return 'kia'
    if key == 'hyundai':
        return 'hyundai'
    if key == 'toyota':
        return 'toyota'
    if key == 'audi':
        return 'audi'

    return key


num_re = re.compile(r'\d+(?:\.\d+)?')


def parse_price(v):
    if v is None:
        return None, None

    if isinstance(v, (int, float, Decimal)):
        d = Decimal(str(v)).quantize(Decimal('0.01'))
        return d, None

    s = str(v).strip().replace('$', '').replace(',', '')
    if not s:
        return None, None

    matches = num_re.findall(s)
    if not matches:
        return None, s

    try:
        d = Decimal(matches[0]).quantize(Decimal('0.01'))
        keep_raw = s if ('-' in s or '+' in s or len(matches) > 1) else None
        return d, keep_raw
    except InvalidOperation:
        return None, s


def sql_text(v):
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def sql_num(v):
    if v is None:
        return 'NULL'
    return str(v)


def parse_year(v):
    if v is None:
        return None

    try:
        year = int(v)
    except Exception:
        t = str(v).strip()
        m = re.search(r'(19|20)\d{2}', t)
        if not m:
            return None
        year = int(m.group(0))

    if year < 1900 or year > 2100:
        return None
    return year


rows_sql = []
for r in range(2, ws.max_row + 1):
    year = parse_year(ws.cell(r, idx['YEAR'] + 1).value)
    make = normalize_make(clean_text(ws.cell(r, idx['MAKE'] + 1).value))
    model = collapse_spaces(clean_text(ws.cell(r, idx['MODEL'] + 1).value))
    ignition = clean_text(ws.cell(r, idx['TYPE_OF_IGNITION'] + 1).value)
    key_type = clean_text(ws.cell(r, idx['TYPE_OF_KEY'] + 1).value)

    if not year or not make or not model or not ignition or not key_type:
        continue

    raw_buttons = ws.cell(r, idx['NO_BUTTONS'] + 1).value
    try:
        buttons = int(raw_buttons) if raw_buttons is not None else 0
    except Exception:
        buttons = 0
    if buttons < 0:
        buttons = 0

    sl_tr, sl_tr_raw = parse_price(ws.cell(r, idx['SL_TRANSMITTER_PRICE '] + 1).value)
    sl_add, sl_add_raw = parse_price(ws.cell(r, idx['SL_ADD_KEY_PRICE'] + 1).value)
    sl_prog, sl_prog_raw = parse_price(ws.cell(r, idx['SL_PROGRAM_PRICE'] + 1).value)

    d_tr, d_tr_raw = parse_price(ws.cell(r, idx['DEALER_TRANSMITTER_PRICE'] + 1).value)
    d_bl, d_bl_raw = parse_price(ws.cell(r, idx['DEALER_EMERGENCY_BLADE_PRICE'] + 1).value)
    d_prog, d_prog_raw = parse_price(ws.cell(r, idx['DEALER_PROGRAM_PRICE'] + 1).value)

    dealer_name = clean_text(ws.cell(r, idx['DEALER_NAME'] + 1).value) or 'Secure Locks Baseline'
    dealer_phone = clean_text(ws.cell(r, idx['DEALER_PHONE'] + 1).value)
    dealer_address = clean_text(ws.cell(r, idx['DEALER_ADDRESS'] + 1).value)

    date_called = ws.cell(r, idx['DEALER_DATE_CALLED'] + 1).value
    if hasattr(date_called, 'date'):
        date_called = date_called.date().isoformat()
    else:
        date_called = clean_text(date_called)

    vin = clean_text(ws.cell(r, idx['DEALER_VIN'] + 1).value)
    part_number = clean_text(ws.cell(r, idx['DEALER_PART_NUMBER'] + 1).value)
    link = clean_text(ws.cell(r, idx['LINK'] + 1).value)

    comments = clean_text(ws.cell(r, idx['COMMENTS'] + 1).value)
    extra = '; '.join(
        [
            f"{label}: {raw}"
            for label, raw in [
                ('SL transmitter raw', sl_tr_raw),
                ('SL add key raw', sl_add_raw),
                ('SL program raw', sl_prog_raw),
                ('Dealer transmitter raw', d_tr_raw),
                ('Dealer blade raw', d_bl_raw),
                ('Dealer program raw', d_prog_raw),
            ]
            if raw
        ]
    )

    if extra:
        comments = f"{comments}; {extra}" if comments else extra

    values = [
        sql_num(year),
        sql_text(make),
        sql_text(model),
        sql_text(ignition),
        sql_text(key_type),
        sql_num(buttons),
        sql_num(sl_tr),
        sql_num(sl_add),
        sql_num(sl_prog),
        sql_num(d_tr),
        sql_num(d_bl),
        sql_num(d_prog),
        sql_text(dealer_name),
        sql_text(dealer_phone),
        sql_text(dealer_address),
        sql_text(vin),
        sql_text(part_number),
        sql_text(date_called),
        sql_text(link),
        sql_text(comments),
    ]

    rows_sql.append('  (' + ', '.join(values) + ')')

sql = f"""-- Generated from SL_Keys-AI.xlsx on 2026-06-12
-- Source rows imported: {len(rows_sql)}

BEGIN;

DELETE FROM current_prices;
DELETE FROM vehicle_configurations;
DELETE FROM models;
DELETE FROM makes;

CREATE TEMP TABLE tmp_seed_raw (
    year integer,
    make text,
    model text,
    type_of_ignition text,
    type_of_key text,
    buttons_count smallint,
    price_secure_locks_akl numeric(10,2),
    price_secure_locks_add_key numeric(10,2),
    price_secure_locks_program_only numeric(10,2),
    price_dealer_transmitter numeric(10,2),
    price_dealer_blade numeric(10,2),
    price_dealer_program numeric(10,2),
    dealer_name text,
    dealer_phone text,
    dealer_address text,
    vin text,
    part_number text,
    date_called text,
    link text,
    comments text
);

INSERT INTO tmp_seed_raw (
    year,
    make,
    model,
    type_of_ignition,
    type_of_key,
    buttons_count,
    price_secure_locks_akl,
    price_secure_locks_add_key,
    price_secure_locks_program_only,
    price_dealer_transmitter,
    price_dealer_blade,
    price_dealer_program,
    dealer_name,
    dealer_phone,
    dealer_address,
    vin,
    part_number,
    date_called,
    link,
    comments
) VALUES
{',\n'.join(rows_sql)};

INSERT INTO makes (name)
SELECT DISTINCT make
FROM tmp_seed_raw
ON CONFLICT (name) DO NOTHING;

INSERT INTO ignition_types (name)
SELECT DISTINCT type_of_ignition
FROM tmp_seed_raw
ON CONFLICT (name) DO NOTHING;

INSERT INTO key_types (name)
SELECT DISTINCT type_of_key
FROM tmp_seed_raw
ON CONFLICT (name) DO NOTHING;

WITH model_candidates AS (
    SELECT
        mk.id AS make_id,
        lower(r.model) AS model_key,
        MIN(r.model) AS model_name
    FROM tmp_seed_raw r
    JOIN makes mk ON mk.name = r.make
    GROUP BY mk.id, lower(r.model)
)
INSERT INTO models (name, make_id)
SELECT mc.model_name, mc.make_id
FROM model_candidates mc
WHERE NOT EXISTS (
    SELECT 1
    FROM models m
    WHERE m.make_id = mc.make_id
        AND lower(m.name) = mc.model_key
);

INSERT INTO dealers (name, phone, address)
SELECT
    r.dealer_name,
    MAX(r.dealer_phone) AS dealer_phone,
    MAX(r.dealer_address) AS dealer_address
FROM tmp_seed_raw r
GROUP BY r.dealer_name
ON CONFLICT (name) DO UPDATE
SET phone = COALESCE(EXCLUDED.phone, dealers.phone),
        address = COALESCE(EXCLUDED.address, dealers.address);

INSERT INTO vehicle_configurations (
    year,
    make_id,
    model_id,
    ignition_type_id,
    key_type_id,
    buttons_count
)
SELECT DISTINCT
    r.year::smallint,
    mk.id,
    md.model_id,
    it.ignition_id,
    kt.key_type_id,
    r.buttons_count::smallint
FROM tmp_seed_raw r
JOIN makes mk ON mk.name = r.make
JOIN LATERAL (
    SELECT MIN(id) AS model_id
    FROM models
    WHERE make_id = mk.id
    AND lower(name) = lower(r.model)
) md ON md.model_id IS NOT NULL
JOIN LATERAL (
    SELECT MIN(id) AS ignition_id
    FROM ignition_types
    WHERE name = r.type_of_ignition
) it ON it.ignition_id IS NOT NULL
JOIN LATERAL (
    SELECT MIN(id) AS key_type_id
    FROM key_types
    WHERE name = r.type_of_key
) kt ON kt.key_type_id IS NOT NULL
WHERE NOT EXISTS (
    SELECT 1
    FROM vehicle_configurations vc
    WHERE vc.year = r.year::smallint
        AND vc.make_id = mk.id
        AND vc.model_id = md.model_id
        AND vc.ignition_type_id = it.ignition_id
        AND vc.key_type_id = kt.key_type_id
        AND vc.buttons_count = r.buttons_count::smallint
);

WITH actor AS (
    SELECT MIN(id) AS user_id
    FROM users
),
prepared_prices AS (
    SELECT
        vc_cfg.vc_id AS vehicle_configuration_id,
        d.id AS dealer_id,
        r.price_secure_locks_akl,
        r.price_secure_locks_add_key,
        r.price_secure_locks_program_only,
        r.price_dealer_transmitter,
        r.price_dealer_program,
        r.price_dealer_blade,
        r.vin,
        r.part_number,
        r.link,
        r.comments,
        CASE
            WHEN r.date_called ~ '^\\d{{4}}-\\d{{2}}-\\d{{2}}$' THEN r.date_called::date
            ELSE NULL
        END AS date_called,
        a.user_id AS created_by_user_id,
        a.user_id AS updated_by_user_id,
        ROW_NUMBER() OVER (
            PARTITION BY vc_cfg.vc_id, d.id
            ORDER BY
                CASE
                    WHEN r.date_called ~ '^\\d{{4}}-\\d{{2}}-\\d{{2}}$' THEN r.date_called::date
                END DESC NULLS LAST
        ) AS rn
    FROM tmp_seed_raw r
    JOIN makes mk ON mk.name = r.make
    JOIN LATERAL (
        SELECT MIN(id) AS model_id
        FROM models
        WHERE make_id = mk.id
            AND lower(name) = lower(r.model)
    ) md ON md.model_id IS NOT NULL
    JOIN LATERAL (
        SELECT MIN(id) AS ignition_id
        FROM ignition_types
        WHERE name = r.type_of_ignition
    ) it ON it.ignition_id IS NOT NULL
    JOIN LATERAL (
        SELECT MIN(id) AS key_type_id
        FROM key_types
        WHERE name = r.type_of_key
    ) kt ON kt.key_type_id IS NOT NULL
    JOIN LATERAL (
        SELECT MIN(id) AS vc_id
        FROM vehicle_configurations vc
        WHERE vc.year = r.year::smallint
            AND vc.make_id = mk.id
            AND vc.model_id = md.model_id
            AND vc.ignition_type_id = it.ignition_id
            AND vc.key_type_id = kt.key_type_id
            AND vc.buttons_count = r.buttons_count::smallint
    ) vc_cfg ON vc_cfg.vc_id IS NOT NULL
    JOIN dealers d ON d.name = r.dealer_name
    JOIN actor a ON a.user_id IS NOT NULL
)
INSERT INTO current_prices (
    vehicle_configuration_id,
    dealer_id,
    price_secure_locks_akl,
    price_secure_locks_add_key,
    price_secure_locks_program_only,
    price_dealer_transmitter,
    price_dealer_program,
    price_dealer_blade,
    vin,
    part_number,
    link,
    comments,
    date_called,
    created_by_user_id,
    updated_by_user_id
)
SELECT
    vehicle_configuration_id,
    dealer_id,
    MAX(price_secure_locks_akl),
    MAX(price_secure_locks_add_key),
    MAX(price_secure_locks_program_only),
    MAX(price_dealer_transmitter),
    MAX(price_dealer_program),
    MAX(price_dealer_blade),
    MAX(vin),
    MAX(part_number),
    MAX(link),
    MAX(comments),
    MAX(date_called),
    MAX(created_by_user_id),
    MAX(updated_by_user_id)
FROM prepared_prices
WHERE rn = 1
GROUP BY vehicle_configuration_id, dealer_id
ON CONFLICT (vehicle_configuration_id, dealer_id) DO NOTHING;

DROP TABLE IF EXISTS tmp_seed_raw;

COMMIT;
"""

OUT.write_text(sql, encoding='utf-8')
print(f'Wrote {OUT} with {len(rows_sql)} rows')
